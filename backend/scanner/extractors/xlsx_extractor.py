"""
Excel文档提取器
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .base import BaseExtractor
from models.schemas import FileInfo, DocumentMetrics


class XlsxExtractor(BaseExtractor):
    """Excel文档提取器"""
    
    def can_handle(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in ['.xlsx', '.xls']
    
    def extract(self, file_path: Path, file_info: FileInfo) -> DocumentMetrics:
        """提取Excel文档指标"""
        metrics = DocumentMetrics()
        
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            
            # Sheet统计
            metrics.sheet_count = len(wb.sheetnames)
            
            total_chars = 0
            total_cells = 0
            merged_count = 0
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # 遍历所有单元格
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            total_cells += 1
                            cell_text = str(cell.value)
                            total_chars += len(cell_text)
                
                # 合并单元格统计(read_only模式下需要特殊处理)
                if hasattr(sheet, 'merged_cells'):
                    merged_count += len(sheet.merged_cells.ranges)
            
            metrics.char_count = total_chars
            metrics.merged_cell_count = merged_count
            
            # 对于Excel,用非空单元格数作为"行数"的替代指标
            metrics.paragraph_count = total_cells
            
            wb.close()
            
        except InvalidFileException:
            file_info.is_corrupted = True
            file_info.parse_success = False
            file_info.parse_error = "文件损坏或格式不正确"
        except Exception as e:
            file_info.parse_success = False
            file_info.parse_error = str(e)
        
        return metrics
    
    def extract_text(self, file_path: Path) -> str:
        """提取文本内容"""
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            texts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    row_texts = []
                    for cell in row:
                        if cell.value is not None:
                            row_texts.append(str(cell.value))
                    if row_texts:
                        texts.append(' '.join(row_texts))
            
            wb.close()
            return '\n'.join(texts)
        except Exception:
            return ""
